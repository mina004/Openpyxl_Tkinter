from __future__ import annotations
from typing import Callable, Dict, List, Optional
from openpyxl import load_workbook


def get_sheet_names(file_path: str) -> List[str]:
    """Return all sheet names using openpyxl (no data load)."""
    # Will raise FileNotFoundError if the path is wrong (as the test expects)
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _row_has_data(values) -> bool:
    """True if at least one cell is non-empty/non-whitespace."""
    for v in values:
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip() != "":
                return True
        else:
            # any non-None non-empty value (numbers, dates, etc.)
            return True
    return False


def get_sheet_row_count(file_path: str, sheet_name: str) -> int:
    """
    Count the number of data rows in the given sheet, EXCLUDING the header row.
    A 'data row' is any row where at least one cell is non-empty.
    """
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        count = 0
        # start from row 2 to exclude header
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _row_has_data(row):
                count += 1
        return count
    finally:
        wb.close()


def process_excel_file(
    file_path: str,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> Dict[str, int]:
    """
    Process all sheets and return {sheet_name: row_count}.
    If progress_callback is provided, call it as (current, total, sheet_name).
    """
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
    finally:
        wb.close()

    results: Dict[str, int] = {}
    total = len(names)

    for i, name in enumerate(names, start=1):
        results[name] = get_sheet_row_count(file_path, name)
        if progress_callback:
            progress_callback(i, total, name)

    return results
