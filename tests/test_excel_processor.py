"""Tests for Excel processing functions.

This module tests the core Excel processing functionality using simple
test cases with pre-created Excel files.
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from xlsx_reader.excel_processor import get_sheet_names, get_sheet_row_count, process_excel_file


def create_test_excel(file_path: str) -> None:
    """Create a simple test Excel file with multiple sheets."""
    workbook = Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Sheet 1: 10 rows
    sheet1 = workbook.create_sheet("Sheet1")
    sheet1["A1"] = "A"
    sheet1["B1"] = "B"
    for i in range(1, 11):
        sheet1[f"A{i + 1}"] = i
        sheet1[f"B{i + 1}"] = f"Row {i}"

    # Sheet 2: 5 rows
    sheet2 = workbook.create_sheet("Sheet2")
    sheet2["X1"] = "X"
    sheet2["Y1"] = "Y"
    for i in range(1, 6):
        sheet2[f"X{i + 1}"] = i
        sheet2[f"Y{i + 1}"] = f"Data {i}"

    # Sheet 3: Empty (0 rows) - just headers
    sheet3 = workbook.create_sheet("EmptySheet")
    sheet3["A1"] = "Col1"
    sheet3["B1"] = "Col2"

    workbook.save(file_path)


class TestExcelProcessor:
    """Test cases for Excel processing functions."""

    @pytest.fixture
    def test_excel_file(self):
        """Create a temporary Excel file for testing."""
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = Path(tmp.name)
        try:
            tmp.close()
            create_test_excel(str(tmp_path))
            yield str(tmp_path)
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_get_sheet_names(self, test_excel_file):
        """Test getting sheet names from Excel file."""
        sheet_names = get_sheet_names(test_excel_file)

        assert isinstance(sheet_names, list)
        assert len(sheet_names) == 3
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names
        assert "EmptySheet" in sheet_names

    def test_get_sheet_row_count_normal_sheet(self, test_excel_file):
        """Test getting row count from a normal sheet."""
        row_count = get_sheet_row_count(test_excel_file, "Sheet1")
        assert row_count == 10

    def test_get_sheet_row_count_small_sheet(self, test_excel_file):
        """Test getting row count from a smaller sheet."""
        row_count = get_sheet_row_count(test_excel_file, "Sheet2")
        assert row_count == 5

    def test_get_sheet_row_count_empty_sheet(self, test_excel_file):
        """Test getting row count from an empty sheet."""
        row_count = get_sheet_row_count(test_excel_file, "EmptySheet")
        assert row_count == 0

    def test_process_excel_file(self, test_excel_file):
        """Test processing entire Excel file."""
        results = process_excel_file(test_excel_file)

        assert isinstance(results, dict)
        assert len(results) == 3
        assert results["Sheet1"] == 10
        assert results["Sheet2"] == 5
        assert results["EmptySheet"] == 0

    def test_process_excel_file_with_callback(self, test_excel_file):
        """Test processing Excel file with progress callback."""
        progress_calls = []

        def progress_callback(current, total, sheet_name):
            progress_calls.append((current, total, sheet_name))

        process_excel_file(test_excel_file, progress_callback)

        # Check that callback was called for each sheet
        assert len(progress_calls) == 3
        assert progress_calls[0] == (0, 3, "Sheet1")
        assert progress_calls[1] == (1, 3, "Sheet2")
        assert progress_calls[2] == (2, 3, "EmptySheet")

    def test_file_not_found(self):
        """Test handling of non-existent file."""
        with pytest.raises(FileNotFoundError):
            get_sheet_names("nonexistent.xlsx")

        with pytest.raises(FileNotFoundError):
            get_sheet_row_count("nonexistent.xlsx", "Sheet1")

        with pytest.raises(FileNotFoundError):
            process_excel_file("nonexistent.xlsx")
